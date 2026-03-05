"""
Report generation service.
Provides PDF and Excel export for transaction lists and financial summaries.
All functions are pure: they receive data and return bytes, no DB access.
"""
import io
from datetime import datetime


def generate_transactions_pdf(transactions: list) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter,
        rightMargin=0.5 * inch, leftMargin=0.5 * inch,
        topMargin=0.75 * inch, bottomMargin=0.75 * inch,
    )
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Transaction Report", styles["Title"]))
    elements.append(Paragraph(
        f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
        styles["Normal"],
    ))
    elements.append(Spacer(1, 0.25 * inch))

    headers = ["Date", "Description", "Category", "Account", "Amount"]
    rows = [headers]
    for t in transactions:
        rows.append([
            t.date.strftime("%Y-%m-%d"),
            t.description[:40],
            t.category or "",
            t.account or "",
            f"${t.amount:+,.2f}",
        ])

    col_widths = [1.0 * inch, 2.5 * inch, 1.2 * inch, 1.2 * inch, 1.1 * inch]
    table = Table(rows, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3b82f6")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e5e7eb")),
        ("ALIGN", (4, 0), (4, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return buffer.read()


def generate_transactions_excel(transactions: list) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    header_fill = PatternFill("solid", fgColor="3B82F6")
    header_font = Font(bold=True, color="FFFFFF")
    for col, header in enumerate(["Date", "Description", "Category", "Account", "Amount"], 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, t in enumerate(transactions, 2):
        ws.cell(row=row_idx, column=1, value=t.date.strftime("%Y-%m-%d"))
        ws.cell(row=row_idx, column=2, value=t.description)
        ws.cell(row=row_idx, column=3, value=t.category or "")
        ws.cell(row=row_idx, column=4, value=t.account or "")
        amount_cell = ws.cell(row=row_idx, column=5, value=t.amount)
        amount_cell.number_format = '#,##0.00'
        amount_cell.font = Font(color="DC2626" if t.amount < 0 else "16A34A")

    for col, width in zip("ABCDE", [14, 38, 18, 18, 14]):
        ws.column_dimensions[col].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def generate_summary_pdf(summary: dict) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter,
        rightMargin=inch, leftMargin=inch,
        topMargin=inch, bottomMargin=inch,
    )
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Financial Summary Report", styles["Title"]))
    elements.append(Paragraph(
        f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
        styles["Normal"],
    ))
    elements.append(Spacer(1, 0.3 * inch))

    kpi_data = [
        ["Metric", "Value"],
        ["Total Revenue", f"${summary.get('total_income', 0):,.2f}"],
        ["Total Expenses", f"${summary.get('total_expenses', 0):,.2f}"],
        ["Net Income", f"${summary.get('net', 0):+,.2f}"],
        ["Transaction Count", str(summary.get("transaction_count", 0))],
    ]
    kpi_table = Table(kpi_data, colWidths=[3 * inch, 2.5 * inch])
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3b82f6")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f9ff")]),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(kpi_table)

    monthly = summary.get("monthly_breakdown", {})
    if monthly:
        elements.append(Spacer(1, 0.3 * inch))
        elements.append(Paragraph("Monthly Breakdown", styles["Heading2"]))
        monthly_rows = [["Month", "Net Cash Flow"]]
        for month, amount in sorted(monthly.items()):
            monthly_rows.append([month, f"${amount:+,.2f}"])
        m_table = Table(monthly_rows, colWidths=[3 * inch, 2.5 * inch])
        m_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6366f1")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        elements.append(m_table)

    doc.build(elements)
    buffer.seek(0)
    return buffer.read()


def generate_tax_report_pdf(tax_data: dict, year: int = None) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter,
        rightMargin=inch, leftMargin=inch,
        topMargin=0.75 * inch, bottomMargin=0.75 * inch,
    )
    styles = getSampleStyleSheet()
    elements = []

    title = f"Tax Summary Report{f' — {year}' if year else ''}"
    elements.append(Paragraph(title, styles["Title"]))
    elements.append(Paragraph(
        f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}  |  Based on IRS Schedule C categories",
        styles["Normal"],
    ))
    elements.append(Spacer(1, 0.2 * inch))

    # Total deductible KPI
    total = tax_data.get("total_deductible", 0)
    kpi_data = [
        ["Total Deductible Expenses", f"${total:,.2f}"],
    ]
    kpi_table = Table(kpi_data, colWidths=[3.5 * inch, 2 * inch])
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#dcfce7")),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#166534")),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 12),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#bbf7d0")),
    ]))
    elements.append(kpi_table)
    elements.append(Spacer(1, 0.25 * inch))

    # By category breakdown
    elements.append(Paragraph("Deductible Expenses by IRS Category", styles["Heading2"]))
    elements.append(Spacer(1, 0.1 * inch))

    by_cat = tax_data.get("by_category", {})
    rows = [["IRS Category", "Deductible Amount", "Note"]]
    for cat, amount in by_cat.items():
        note = "50% deductible" if cat == "Meals (50%)" else ""
        rows.append([cat, f"${amount:,.2f}", note])

    cat_table = Table(rows, colWidths=[2.5 * inch, 1.8 * inch, 1.8 * inch], repeatRows=1)
    cat_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4f46e5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f3ff")]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e5e7eb")),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(cat_table)
    elements.append(Spacer(1, 0.3 * inch))

    disclaimer = ParagraphStyle("disclaimer", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#6b7280"))
    elements.append(Paragraph(
        "⚠ For informational purposes only. This report does not constitute tax advice. "
        "Please consult a qualified tax professional before filing.",
        disclaimer,
    ))

    doc.build(elements)
    buffer.seek(0)
    return buffer.read()


def generate_summary_excel(summary: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    blue_fill = PatternFill("solid", fgColor="3B82F6")
    bold_white = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(["Metric", "Value"], 1):
        c = ws.cell(row=1, column=col, value=header)
        c.fill = blue_fill
        c.font = bold_white
        c.alignment = Alignment(horizontal="center")

    kpis = [
        ("Total Revenue", summary.get("total_income", 0)),
        ("Total Expenses", summary.get("total_expenses", 0)),
        ("Net Income", summary.get("net", 0)),
        ("Transaction Count", summary.get("transaction_count", 0)),
    ]
    for row_idx, (label, value) in enumerate(kpis, 2):
        ws.cell(row=row_idx, column=1, value=label)
        v_cell = ws.cell(row=row_idx, column=2, value=value)
        if isinstance(value, float):
            v_cell.number_format = '#,##0.00'

    monthly = summary.get("monthly_breakdown", {})
    if monthly:
        start_row = len(kpis) + 4
        ws.cell(row=start_row - 1, column=1, value="Monthly Breakdown").font = Font(bold=True)
        purple_fill = PatternFill("solid", fgColor="6366F1")
        for col, header in enumerate(["Month", "Net Cash Flow"], 1):
            c = ws.cell(row=start_row, column=col, value=header)
            c.fill = purple_fill
            c.font = bold_white
        for row_idx, (month, amount) in enumerate(sorted(monthly.items()), start_row + 1):
            ws.cell(row=row_idx, column=1, value=month)
            mc = ws.cell(row=row_idx, column=2, value=amount)
            mc.number_format = '#,##0.00'

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
